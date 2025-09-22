import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI
from openai.types.chat import ChatCompletionChunk
from dotenv import load_dotenv
import time
import argparse
import logging
from tqdm import tqdm
import re
from datetime import datetime
import shutil
import json
import signal
import asyncio
import sys

# Set up logging
log_filename = f"excel_translation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Create cache directory if it doesn't exist
cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "translation_cache")
os.makedirs(cache_dir, exist_ok=True)

# Initialize OpenAI client with proper timeout settings
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    timeout=90.0  # Increased timeout to 90 seconds
)

def is_chinese(text):
    """
    Check if text contains Chinese characters.
    
    Args:
        text (str): Text to check
        
    Returns:
        bool: True if text contains Chinese characters
    """
    if not isinstance(text, str):
        return False
    
    # Chinese Unicode ranges (expanded):
    # Basic Chinese: 4E00-9FFF
    # Extended Chinese A: 3400-4DBF
    # Extended Chinese B: 20000-2A6DF
    # Extended Chinese C: 2A700-2B73F
    # Extended Chinese D: 2B740-2B81F
    # Extended Chinese E: 2B820-2CEAF
    # CJK Radicals: 2F00-2FDF
    # CJK Symbols and Punctuation: 3000-303F
    chinese_pattern = re.compile(r'[\u4E00-\u9FFF\u3400-\u4DBF\u20000-\u2A6DF\u2A700-\u2B73F\u2B740-\u2B81F\u2B820-\u2CEAF\u2F00-\u2FDF\u3000-\u303F]')
    return bool(chinese_pattern.search(text))

def get_cache_key(text):
    """Generate a unique cache key for the text."""
    import hashlib
    return hashlib.md5(text.encode('utf-8')).hexdigest()

def load_translation_cache():
    """Load translation cache from disk."""
    cache = {}
    try:
        cache_file = os.path.join(cache_dir, "translation_cache.json")
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            logger.info(f"Loaded {len(cache)} translations from cache")
    except Exception as e:
        logger.warning(f"Failed to load translation cache: {e}")
    return cache

def save_translation_cache(cache):
    """Save translation cache to disk."""
    try:
        cache_file = os.path.join(cache_dir, "translation_cache.json")
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        logger.debug(f"Saved {len(cache)} translations to cache")
    except Exception as e:
        logger.warning(f"Failed to save translation cache: {e}")

async def stream_translation(text, context="questionnaire"):
    """
    Stream a translation from the OpenAI API using the streaming interface.
    
    Args:
        text (str): Text to translate
        context (str): Context to help with translation
        
    Returns:
        str: Translated text
    """
    messages = [
        {"role": "system", "content": "You are a direct Chinese-to-English translator. ONLY translate the exact Chinese text provided. Do not add ANY formatting, explanations, or extra words. Keep punctuation, numbers, and special characters exactly as they appear in the original. Your job is ONLY literal translation."},
        {"role": "user", "content": f"Translate this text to English (provide ONLY the direct translation without ANY additional text): {text}"}
    ]
    
    try:
        # Creating a new client for each streaming call to avoid any potential issues
        # with concurrent streaming requests
        stream = await asyncio.to_thread(
            lambda: client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                temperature=0.1,
                stream=True
            )
        )
        
        collected_chunks = []
        for chunk in stream:
            collected_chunks.append(chunk)
        
        content_list = []
        for chunk in collected_chunks:
            if hasattr(chunk.choices[0].delta, 'content') and chunk.choices[0].delta.content is not None:
                content_list.append(chunk.choices[0].delta.content)
        
        full_content = ''.join(content_list)
        return full_content.strip()
        
    except Exception as e:
        logger.error(f"Streaming translation error for text '{text[:50]}...': {e}")
        return None

async def batch_translate_texts_async(texts, context="running apps", max_retries=5, base_delay=2):
    """
    Asynchronously translate multiple texts using streaming API.
    
    Args:
        texts (list): List of texts to translate
        context (str): Context to help with translation
        max_retries (int): Maximum number of retries
        base_delay (float): Base delay between retries in seconds
        
    Returns:
        dict: Dictionary mapping original texts to translations
    """
    # Load cache
    if not hasattr(batch_translate_texts_async, "cache"):
        batch_translate_texts_async.cache = load_translation_cache()
    
    # Filter out texts that don't need translation
    texts_to_translate = []
    results = {}
    
    for text in texts:
        if not isinstance(text, str) or not text.strip():
            results[text] = text
            continue
            
        # Check if text contains Chinese characters
        if not is_chinese(text):
            results[text] = text
            continue
            
        # Check cache
        cache_key = get_cache_key(text)
        if cache_key in batch_translate_texts_async.cache:
            results[text] = batch_translate_texts_async.cache[cache_key]
            continue
            
        texts_to_translate.append(text)
    
    # If no texts need translation, return results
    if not texts_to_translate:
        return results
    
    # Process each text with streaming API and retry logic
    for text in texts_to_translate:
        for attempt in range(max_retries + 1):
            try:
                translated_text = await stream_translation(text, context)
                
                if translated_text:
                    results[text] = translated_text
                    cache_key = get_cache_key(text)
                    batch_translate_texts_async.cache[cache_key] = translated_text
                    # Save cache periodically
                    save_translation_cache(batch_translate_texts_async.cache)
                    break  # Success, exit retry loop
                else:
                    # If translation returned None, retry
                    if attempt < max_retries:
                        delay = base_delay * (2 ** attempt)  # Exponential backoff
                        logger.warning(f"Attempt {attempt+1}/{max_retries+1} failed for text '{text[:30]}...'. Retrying in {delay} seconds...")
                        await asyncio.sleep(delay)
                    else:
                        logger.error(f"Translation failed after {max_retries+1} attempts for text '{text[:30]}...'")
                        results[text] = text  # Use original text as fallback
            except Exception as e:
                if attempt < max_retries:
                    delay = base_delay * (2 ** attempt)  # Exponential backoff
                    logger.warning(f"Attempt {attempt+1}/{max_retries+1} failed for text '{text[:30]}...': {e}")
                    logger.warning(f"Retrying in {delay} seconds...")
                    await asyncio.sleep(delay)
                else:
                    logger.error(f"Translation failed after {max_retries+1} attempts for text '{text[:30]}...': {e}")
                    results[text] = text  # Use original text as fallback
    
    return results

def batch_translate_texts(texts, context="none", max_retries=5, base_delay=2):
    """
    Translate multiple texts at once using streaming API (synchronous wrapper).
    
    Args:
        texts (list): List of texts to translate
        context (str): Context to help with translation
        max_retries (int): Maximum number of retries
        base_delay (float): Base delay between retries in seconds
        
    Returns:
        dict: Dictionary mapping original texts to translations
    """
    # If asyncio is already running, we need to use the running loop
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            # Create a new loop for this thread if the current one is running
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
    except RuntimeError:
        # If there is no event loop, create a new one
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    
    return loop.run_until_complete(batch_translate_texts_async(texts, context, max_retries, base_delay))

def parse_batch_translations(content, expected_count):
    """
    Parse the batch translation response into individual translations.
    
    Args:
        content (str): The translated content from the API
        expected_count (int): The expected number of translations
        
    Returns:
        list: List of translated texts
    """
    translations = []
    
    # Try to parse numbered format (1. Translation\n2. Translation)
    import re
    pattern = r'\d+\.\s*(.+?)(?=\n\d+\.|$)'  # Match numbered translations
    matches = re.findall(pattern, content, re.DOTALL)
    
    if matches and len(matches) == expected_count:
        return [match.strip() for match in matches]
    
    # Try splitting by separators if numbering fails
    if "---" in content:
        parts = content.split("---")
        translations = [part.strip() for part in parts if part.strip()]
    else:
        # Split by newlines as last resort
        translations = [line.strip() for line in content.split("\n") if line.strip()]
        
        # Try to clean up numbering if present
        cleaned = []
        for t in translations:
            if re.match(r'^\d+\.\s*', t):  # If line starts with a number and period
                t = re.sub(r'^\d+\.\s*', '', t)  # Remove the numbering
            cleaned.append(t)
        translations = cleaned
    
    # Ensure we don't return more translations than expected
    return translations[:expected_count]

def translate_text(text, context="running apps", max_retries=5, base_delay=2):
    """
    Translate a single text from Chinese to English.
    This is a wrapper around batch_translate_texts for backward compatibility.
    
    Args:
        text (str): Text to translate
        context (str): Context to help with translation
        max_retries (int): Maximum number of retries
        base_delay (float): Base delay between retries in seconds
        
    Returns:
        str: Translated text
    """
    # Skip translation if text is not a string or doesn't contain Chinese
    if not isinstance(text, str) or not text.strip() or not is_chinese(text):
        return text
        
    # Check cache first
    if not hasattr(translate_text, "cache"):
        translate_text.cache = load_translation_cache()
    
    cache_key = get_cache_key(text)
    if cache_key in translate_text.cache:
        return translate_text.cache[cache_key]
    
    # Use batch translate for actual translation (which uses streaming API)
    results = batch_translate_texts([text], context, max_retries, base_delay)
    return results.get(text, text)

def create_backup(file_path):
    """
    Create a backup of the original file.
    
    Args:
        file_path (str): Path to the file to backup
        
    Returns:
        str: Path to the backup file
    """
    backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    try:
        shutil.copy2(file_path, backup_path)
        logger.info(f"Created backup at {backup_path}")
        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup: {e}")
        return None

def translate_excel_file(input_file, output_file, context="none", batch_size=5, create_backup_file=True, max_retries=5, save_interval=20):
    """
    Translate an Excel file from Chinese to English while preserving exact cell positions.
    
    Args:
        input_file (str): Path to input Excel file
        output_file (str): Path to output Excel file
        context (str): Context to help with translation
        batch_size (int): Number of cells to translate in one batch
        create_backup_file (bool): Whether to create a backup of the original file
        max_retries (int): Maximum number of retries for failed translations
        save_interval (int): Save output file every N cells
    """
    logger.info(f"Starting translation of {input_file}")
    start_time = time.time()
    
    # Create backup if requested
    if create_backup_file:
        backup_file = create_backup(input_file)
        if not backup_file:
            logger.warning("Failed to create backup, proceeding without backup")
    
    # First make a direct copy of the input file to the output file, preserving all cells
    try:
        shutil.copy2(input_file, output_file)
        logger.info(f"Created initial copy of workbook as {output_file}")
    except Exception as e:
        logger.error(f"Error creating initial copy: {e}")
        return
    
    # Now load the copied file for translation
    try:
        workbook = openpyxl.load_workbook(output_file, data_only=False)
        sheet_names = workbook.sheetnames
        logger.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return
    
    # Initialize the cache for batch translations
    if not hasattr(batch_translate_texts, "cache"):
        batch_translate_texts.cache = load_translation_cache()
    
    # Set up signal handler for graceful interruption
    setup_signal_handler(workbook, output_file)
    
    # Collect cells that need translation directly from the output workbook
    all_cells = []
    formula_cells = []
    
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip() and is_chinese(cell.value):
                    if cell.value.startswith('='):  # Formula cell
                        formula_cells.append((sheet_name, cell))
                    else:  # Regular cell
                        all_cells.append((sheet_name, cell))
    
    total_cells = len(all_cells) + len(formula_cells)
    logger.info(f"Found {total_cells} cells to translate across all sheets")
    logger.info(f"Regular cells: {len(all_cells)}, Formula cells: {len(formula_cells)}")
    
    # Create a master progress bar
    master_pbar = tqdm(total=total_cells, desc="Total Progress", position=0)
    
    # Process regular cells in batches
    cells_processed = 0
    cells_translated = 0
    
    for i in range(0, len(all_cells), batch_size):
        # Get a batch of cells
        batch = all_cells[i:i+batch_size]
        cells_processed += len(batch)
        
        # Save workbook periodically
        if cells_processed % save_interval == 0:
            try:
                logger.info(f"Saving progress after {cells_processed} cells...")
                workbook.save(output_file)
            except Exception as e:
                logger.error(f"Error saving intermediate progress: {e}")
        
        # Prepare for batch translation
        texts_to_translate = []
        cell_mapping = []
        
        for sheet_name, cell in batch:
            texts_to_translate.append(cell.value)
            cell_mapping.append((sheet_name, cell))
        
        # Translate the batch
        translations = batch_translate_texts(texts_to_translate, context, max_retries)
        
        # Update cells with translations - keeping exact cell positions and preserving coordinate information
        for idx, (original_text, translation) in enumerate(zip(texts_to_translate, translations.values())):
            if original_text != translation:
                sheet_name, cell = cell_mapping[idx]
                # Get and log precise cell information before translation
                coordinate = cell.coordinate
                row, column = coordinate[0], int(coordinate[1:]) if coordinate[1:].isdigit() else coordinate[1:]
                logger.info(f"Translating cell {sheet_name}!{coordinate} (row {row}, col {column})")
                
                # Update the cell directly - no need to re-get it since we have the reference
                cell.value = translation
                cells_translated += 1
                logger.info(f"Cell {sheet_name}!{coordinate} translated from '{original_text}' to '{translation}'")
            
            # Update progress bar
            master_pbar.update(1)
    
    # Process formula cells one by one
    for sheet_name, cell in formula_cells:
        original_formula = cell.value
        cell_address = f"{sheet_name}!{cell.coordinate}"
        
        logger.debug(f"Processing formula cell {cell_address}: '{original_formula}'")
        
        # Extract text strings from the formula
        string_pattern = re.compile(r'"([^"]*)"')
        strings = string_pattern.findall(original_formula)
        
        # Filter strings that need translation
        chinese_strings = [s for s in strings if is_chinese(s)]
        
        if chinese_strings:
            # Translate the strings
            translations = batch_translate_texts(chinese_strings, context, max_retries)
            
            # Replace strings in the formula
            translated_formula = original_formula
            for string in chinese_strings:
                if string in translations:
                    translated_string = translations[string]
                    if translated_string != string:
                        translated_formula = translated_formula.replace(f'"{string}"', f'"{translated_string}"')
            
            # Update cell if formula changed
            if translated_formula != original_formula:
                cell.value = translated_formula
                cells_translated += 1
                logger.debug(f"Formula cell {cell_address} translated: '{original_formula}' -> '{translated_formula}'")
        
        # Update progress bar
        master_pbar.update(1)
        
        # Save periodically for formula cells too
        cells_processed += 1
        if cells_processed % save_interval == 0:
            try:
                logger.info(f"Saving progress after {cells_processed} cells...")
                workbook.save(output_file)
            except Exception as e:
                logger.error(f"Error saving intermediate progress: {e}")
    
    # Close the progress bar
    master_pbar.close()
    
    # Save the final result
    try:
        workbook.save(output_file)
        end_time = time.time()
        elapsed_time = end_time - start_time
        logger.info(f"Translation completed. Translated {cells_translated} of {total_cells} cells in {elapsed_time:.2f} seconds")
        logger.info(f"Saved to {output_file}")
    except Exception as e:
        logger.error(f"Error saving translated Excel file: {e}")

def setup_signal_handler(workbook, output_file):
    """Set up signal handler to save workbook on interrupt"""
    def signal_handler(sig, frame):
        print("\nSaving workbook before exiting...")
        try:
            workbook.save(output_file)
            print(f"Saved to {output_file}")
        except Exception as e:
            print(f"Error saving workbook: {e}")
        finally:
            # Save translation cache
            if hasattr(batch_translate_texts, "cache"):
                save_translation_cache(batch_translate_texts.cache)
                print("Saved translation cache")
            sys.exit(0)
    
    signal.signal(signal.SIGINT, signal_handler)

def main():
    """
    Main function for CLI usage.
    """
    parser = argparse.ArgumentParser(
        description='Translate Excel files from Chinese to English while preserving formatting'
    )
    parser.add_argument('--input', '-i', required=True, help='Input Excel file')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file')
    parser.add_argument('--context', '-c', default='none', help='Translation context (optional, default: none)')
    parser.add_argument('--batch-size', '-b', type=int, default=5, help='Batch size for translation (default: 5)')
    parser.add_argument('--no-backup', action='store_true', help='Skip creating a backup file')
    parser.add_argument('--max-retries', '-r', type=int, default=5, help='Maximum number of retries for failed translations (default: 5)')
    parser.add_argument('--save-interval', '-s', type=int, default=20, help='Save output file every N cells (default: 20)')
    parser.add_argument('--clear-cache', action='store_true', help='Clear the translation cache before starting')
    
    args = parser.parse_args()
    
    # Clean up file paths - remove quotes if present
    input_file = args.input.strip('"\'')
    output_file = args.output.strip('"\'')
    
    # Pre-load translation cache or clear it if requested
    if args.clear_cache:
        logger.info("Clearing translation cache as requested")
        batch_translate_texts.cache = {}
        save_translation_cache({})
    elif not hasattr(batch_translate_texts, "cache"):
        batch_translate_texts.cache = load_translation_cache()
    # Check if output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"Created output directory: {output_dir}")
    
    translate_excel_file(
        input_file=input_file, 
        output_file=output_file, 
        context=args.context, 
        batch_size=args.batch_size,
        create_backup_file=not args.no_backup
    )

if __name__ == "__main__":
    main() 