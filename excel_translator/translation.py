"""Core Excel translation functionality."""

import os
import re
import time
import shutil
import signal
import asyncio
from datetime import datetime
from typing import List, Tuple, Optional, Dict
import logging

import openpyxl
from tqdm import tqdm

from .providers import OpenAIProvider, BaseProvider
from .utils import is_chinese, get_cache_key, load_cache, save_cache

logger = logging.getLogger(__name__)


class ExcelTranslator:
    """Excel file translator that preserves formatting."""

    def __init__(
        self,
        provider: BaseProvider,
        cache_dir: str = "translation_cache",
        batch_size: int = 5,
        max_retries: int = 5,
        save_interval: int = 20
    ):
        """Initialize the Excel translator.

        Args:
            provider: Translation provider to use
            cache_dir: Directory for translation cache
            batch_size: Number of cells to translate per batch
            max_retries: Maximum number of retries for failed translations
            save_interval: Save progress every N cells
        """
        self.provider = provider
        self.cache_dir = cache_dir
        self.batch_size = batch_size
        self.max_retries = max_retries
        self.save_interval = save_interval

        # Ensure cache directory exists
        os.makedirs(cache_dir, exist_ok=True)
        self.cache_file = os.path.join(cache_dir, "translation_cache.json")
        self.cache = load_cache(self.cache_file)

    def create_backup(self, file_path: str) -> Optional[str]:
        """Create a backup of the original file.

        Args:
            file_path: Path to the file to backup

        Returns:
            Path to the backup file or None if failed
        """
        backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        try:
            shutil.copy2(file_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
            return backup_path
        except Exception as e:
            logger.error(f"Failed to create backup: {e}")
            return None

    def setup_signal_handler(self, workbook: openpyxl.Workbook, output_file: str) -> None:
        """Set up signal handler to save workbook on interrupt.

        Args:
            workbook: Excel workbook to save
            output_file: Output file path
        """
        def signal_handler(sig, frame):
            print("\nSaving workbook before exiting...")
            try:
                workbook.save(output_file)
                print(f"Saved to {output_file}")
            except Exception as e:
                print(f"Error saving workbook: {e}")
            finally:
                save_cache(self.cache, self.cache_file)
                print("Saved translation cache")
                exit(0)

        signal.signal(signal.SIGINT, signal_handler)

    async def translate_cells_batch(
        self,
        cells: List[Tuple[str, any]],
        source_lang: str = "zh",
        target_lang: str = "en",
        context: str = "spreadsheet"
    ) -> Dict[str, str]:
        """Translate a batch of cells.

        Args:
            cells: List of (sheet_name, cell) tuples
            source_lang: Source language code
            target_lang: Target language code
            context: Translation context

        Returns:
            Dictionary mapping original texts to translations
        """
        texts_to_translate = []
        text_cache_map = {}

        for sheet_name, cell in cells:
            text = cell.value
            if not isinstance(text, str) or not text.strip():
                continue

            if not is_chinese(text):
                continue

            # Check cache
            cache_key = get_cache_key(text)
            if cache_key in self.cache:
                text_cache_map[text] = self.cache[cache_key]
                continue

            texts_to_translate.append(text)

        # If all texts are cached, return cache results
        if not texts_to_translate:
            return text_cache_map

        # Translate uncached texts
        for attempt in range(self.max_retries + 1):
            try:
                translations = await self.provider.translate_batch(
                    texts_to_translate,
                    source_lang,
                    target_lang,
                    context
                )

                # Update cache with new translations
                for text, translation in translations.items():
                    if text != translation:
                        cache_key = get_cache_key(text)
                        self.cache[cache_key] = translation
                        text_cache_map[text] = translation
                    else:
                        text_cache_map[text] = text

                # Save cache periodically
                save_cache(self.cache, self.cache_file)
                break

            except Exception as e:
                if attempt < self.max_retries:
                    delay = 2 ** attempt  # Exponential backoff
                    logger.warning(f"Batch translation attempt {attempt+1} failed: {e}. Retrying in {delay}s...")
                    await asyncio.sleep(delay)
                else:
                    logger.error(f"Batch translation failed after {self.max_retries+1} attempts: {e}")
                    # Use original texts as fallback
                    for text in texts_to_translate:
                        text_cache_map[text] = text

        return text_cache_map

    async def translate_file(
        self,
        input_file: str,
        output_file: str,
        source_lang: str = "zh",
        target_lang: str = "en",
        context: str = "spreadsheet",
        create_backup_file: bool = True
    ) -> None:
        """Translate an Excel file while preserving formatting.

        Args:
            input_file: Path to input Excel file
            output_file: Path to output Excel file
            source_lang: Source language code
            target_lang: Target language code
            context: Translation context
            create_backup_file: Whether to create a backup
        """
        logger.info(f"Starting translation of {input_file}")
        start_time = time.time()

        # Create backup if requested
        if create_backup_file:
            backup_file = self.create_backup(input_file)
            if not backup_file:
                logger.warning("Failed to create backup, proceeding without backup")

        # Create initial copy
        try:
            shutil.copy2(input_file, output_file)
            logger.info(f"Created initial copy of workbook as {output_file}")
        except Exception as e:
            logger.error(f"Error creating initial copy: {e}")
            return

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(output_file, data_only=False)
            sheet_names = workbook.sheetnames
            logger.info(f"Found {len(sheet_names)} sheets: {', '.join(sheet_names)}")
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return

        # Set up signal handler
        self.setup_signal_handler(workbook, output_file)

        # Collect cells that need translation
        all_cells = []
        formula_cells = []

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip() and is_chinese(cell.value):
                        if cell.value.startswith('='):  # Formula cell
                            formula_cells.append((sheet_name, cell))
                        else:  # Regular cell
                            all_cells.append((sheet_name, cell))

        total_cells = len(all_cells) + len(formula_cells)
        logger.info(f"Found {total_cells} cells to translate across all sheets")
        logger.info(f"Regular cells: {len(all_cells)}, Formula cells: {len(formula_cells)}")

        # Progress tracking
        master_pbar = tqdm(total=total_cells, desc="Total Progress", position=0)
        cells_processed = 0
        cells_translated = 0

        # Process regular cells in batches
        for i in range(0, len(all_cells), self.batch_size):
            batch = all_cells[i:i+self.batch_size]
            cells_processed += len(batch)

            # Save workbook periodically
            if cells_processed % self.save_interval == 0:
                try:
                    logger.info(f"Saving progress after {cells_processed} cells...")
                    workbook.save(output_file)
                except Exception as e:
                    logger.error(f"Error saving intermediate progress: {e}")

            # Translate batch
            translations = await self.translate_cells_batch(
                batch, source_lang, target_lang, context
            )

            # Update cells with translations
            for sheet_name, cell in batch:
                original_text = cell.value
                if original_text in translations:
                    translation = translations[original_text]
                    if original_text != translation:
                        coordinate = cell.coordinate
                        logger.info(f"Translating cell {sheet_name}!{coordinate}")
                        cell.value = translation
                        cells_translated += 1
                        logger.info(f"Cell {sheet_name}!{coordinate} translated from '{original_text}' to '{translation}'")

                master_pbar.update(1)

        # Process formula cells
        for sheet_name, cell in formula_cells:
            original_formula = cell.value
            cell_address = f"{sheet_name}!{cell.coordinate}"

            logger.debug(f"Processing formula cell {cell_address}: '{original_formula}'")

            # Extract text strings from formula
            string_pattern = re.compile(r'"([^"]*)"')
            strings = string_pattern.findall(original_formula)

            # Filter strings that need translation
            chinese_strings = [s for s in strings if is_chinese(s)]

            if chinese_strings:
                # Translate the strings
                translations = await self.translate_cells_batch(
                    [(sheet_name, type('MockCell', (), {'value': s})()) for s in chinese_strings],
                    source_lang, target_lang, context
                )

                # Replace strings in formula
                translated_formula = original_formula
                for string in chinese_strings:
                    if string in translations:
                        translated_string = translations[string]
                        if translated_string != string:
                            translated_formula = translated_formula.replace(f'"{string}"', f'"{translated_string}"')

                # Update cell if formula changed
                if translated_formula != original_formula:
                    cell.value = translated_formula
                    cells_translated += 1
                    logger.debug(f"Formula cell {cell_address} translated: '{original_formula}' -> '{translated_formula}'")

            master_pbar.update(1)
            cells_processed += 1

            # Save periodically for formula cells too
            if cells_processed % self.save_interval == 0:
                try:
                    logger.info(f"Saving progress after {cells_processed} cells...")
                    workbook.save(output_file)
                except Exception as e:
                    logger.error(f"Error saving intermediate progress: {e}")

        # Close progress bar
        master_pbar.close()

        # Save final result
        try:
            workbook.save(output_file)
            end_time = time.time()
            elapsed_time = end_time - start_time
            logger.info(f"Translation completed. Translated {cells_translated} of {total_cells} cells in {elapsed_time:.2f} seconds")
            logger.info(f"Saved to {output_file}")
        except Exception as e:
            logger.error(f"Error saving translated Excel file: {e}")

    def translate_file_sync(
        self,
        input_file: str,
        output_file: str,
        source_lang: str = "zh",
        target_lang: str = "en",
        context: str = "spreadsheet",
        create_backup_file: bool = True
    ) -> None:
        """Synchronous wrapper for translate_file.

        Args:
            input_file: Path to input Excel file
            output_file: Path to output Excel file
            source_lang: Source language code
            target_lang: Target language code
            context: Translation context
            create_backup_file: Whether to create a backup
        """
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

        loop.run_until_complete(
            self.translate_file(
                input_file, output_file, source_lang, target_lang, context, create_backup_file
            )
        )